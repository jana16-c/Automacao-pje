from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal

from openpyxl.workbook.workbook import Workbook

from pje_automation.domain.models import HistoricalSeries, HistoricalValue, Record, RecordSource, WorkbookPreview
from pje_automation.excel.normalization import (
    normalize_competencia,
    normalize_cpf,
    normalize_date,
    normalize_header,
    normalize_name_key,
    normalize_registration,
    parse_decimal,
)


MAIN_HEADER_ALIASES = {
    "matricula": {"matricula", "registro", "codigo", "codigo empregado"},
    "nome": {"nome", "reclamante", "empregado", "nome reclamante"},
    "cpf": {"cpf", "cpf reclamante", "documento", "documento fiscal"},
    "data_admissao": {"admissao", "data admissao", "dt admissao"},
    "data_demissao": {"demissao", "data demissao", "dt demissao", "data final"},
    "processo": {"processo", "numero processo", "numero do processo"},
    "historico_nome": {"historico nome", "historico salarial", "nome historico", "verba historico"},
}

HISTORY_HEADER_ALIASES = {
    "cpf": MAIN_HEADER_ALIASES["cpf"],
    "matricula": MAIN_HEADER_ALIASES["matricula"],
    "historico_nome": MAIN_HEADER_ALIASES["historico_nome"],
    "nome": MAIN_HEADER_ALIASES["nome"] | {"funcionario"},
    "competencia": {"competencia", "mes ano", "mes/ano", "referencia", "periodo"},
    "valor": {"valor", "valor historico", "valor base", "base informada"},
}


@dataclass(slots=True)
class SheetMapping:
    header_row: int
    columns: dict[str, int]


@dataclass(slots=True)
class HistoryIndex:
    by_cpf: dict[str, dict[str, HistoricalSeries]]
    by_matricula: dict[str, dict[str, HistoricalSeries]]
    by_nome: dict[str, dict[str, HistoricalSeries]]


def build_preview(workbook: Workbook, history_workbook: Workbook | None = None, limit: int = 20) -> WorkbookPreview:
    valid_records: list[Record] = []
    invalid_rows: list[str] = []
    ambiguous_rows: list[str] = []
    history_index, history_invalid_rows = collect_history_series(workbook, history_workbook=history_workbook)
    invalid_rows.extend(history_invalid_rows)

    for worksheet in workbook.worksheets:
        mapping = detect_mapping(
            worksheet.iter_rows(values_only=True),
            aliases=MAIN_HEADER_ALIASES,
            required_fields=("nome", "cpf", "data_demissao"),
        )
        if not mapping:
            continue

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=mapping.header_row + 1, values_only=True),
            start=mapping.header_row + 1,
        ):
            nome = to_optional_str(cell_value(row, mapping.columns.get("nome")))
            if not nome:
                continue

            cpf = normalize_cpf(cell_value(row, mapping.columns.get("cpf")))
            matricula = normalize_registration(cell_value(row, mapping.columns.get("matricula")))
            data_demissao = normalize_date(cell_value(row, mapping.columns.get("data_demissao")))
            if not cpf or not data_demissao:
                invalid_rows.append(f"{worksheet.title}:{row_index}")
                continue

            record_id = cpf or f"{worksheet.title}-{row_index}"
            valid_records.append(
                Record(
                    record_id=record_id,
                    nome=nome,
                    cpf=cpf,
                    data_admissao=normalize_date(cell_value(row, mapping.columns.get("data_admissao"))),
                    data_demissao=data_demissao,
                    processo=to_optional_str(cell_value(row, mapping.columns.get("processo"))),
                    historicos=resolve_historicos(
                        cpf=cpf,
                        matricula=matricula,
                        nome=nome,
                        historico_nome=to_optional_str(cell_value(row, mapping.columns.get("historico_nome"))),
                        history_index=history_index,
                    ),
                    source=RecordSource(sheet=worksheet.title, row=row_index),
                )
            )
            if len(valid_records) >= limit:
                break
        if len(valid_records) >= limit:
            break

    if not valid_records:
        ambiguous_rows.append("Nenhuma linha elegivel foi encontrada com nome, CPF e data de demissao.")

    return WorkbookPreview(
        valid_records=valid_records,
        invalid_rows=invalid_rows,
        ambiguous_rows=ambiguous_rows,
        sheet_names=workbook.sheetnames,
    )


def detect_mapping(
    rows: Iterable[tuple[object, ...]],
    aliases: dict[str, set[str]],
    required_fields: tuple[str, ...],
) -> SheetMapping | None:
    for row_index, row in enumerate(rows, start=1):
        columns: dict[str, int] = {}
        for index, value in enumerate(row):
            header = normalize_header(value)
            if not header:
                continue
            for field, accepted_headers in aliases.items():
                if header in accepted_headers:
                    columns[field] = index
        if all(field in columns for field in required_fields):
            return SheetMapping(header_row=row_index, columns=columns)
        if row_index >= 15:
            break
    return None


def collect_history_series(workbook: Workbook, history_workbook: Workbook | None = None) -> tuple[HistoryIndex, list[str]]:
    history_index = HistoryIndex(by_cpf={}, by_matricula={}, by_nome={})
    invalid_rows: list[str] = []

    for source_workbook in history_sources(workbook, history_workbook):
        invalid_rows.extend(collect_structured_history_series(source_workbook, history_index))
        invalid_rows.extend(collect_companion_history_series(source_workbook, history_index))

    sort_history_values(history_index)
    return history_index, invalid_rows


def collect_structured_history_series(workbook: Workbook, history_index: HistoryIndex) -> list[str]:
    invalid_rows: list[str] = []
    for worksheet in workbook.worksheets:
        mapping = detect_mapping(
            worksheet.iter_rows(values_only=True),
            aliases=HISTORY_HEADER_ALIASES,
            required_fields=("cpf", "historico_nome", "competencia", "valor"),
        )
        if not mapping:
            continue

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=mapping.header_row + 1, values_only=True),
            start=mapping.header_row + 1,
        ):
            cpf = normalize_cpf(cell_value(row, mapping.columns.get("cpf")))
            historico_nome = to_optional_str(cell_value(row, mapping.columns.get("historico_nome")))
            competencia = normalize_competencia(cell_value(row, mapping.columns.get("competencia")))
            raw_valor = cell_value(row, mapping.columns.get("valor"))

            if not any((cpf, historico_nome, competencia, raw_valor)):
                continue
            if not cpf or not historico_nome or not competencia:
                invalid_rows.append(f"{worksheet.title}:{row_index}")
                continue
            if not has_meaningful_history_value(raw_valor):
                continue

            try:
                valor = parse_decimal(raw_valor)
            except ValueError:
                invalid_rows.append(f"{worksheet.title}:{row_index}")
                continue

            add_history_value(history_index.by_cpf, cpf, historico_nome, competencia, valor)

    return invalid_rows


def collect_companion_history_series(workbook: Workbook, history_index: HistoryIndex) -> list[str]:
    invalid_rows: list[str] = []
    for worksheet in workbook.worksheets:
        mapping = detect_mapping(
            worksheet.iter_rows(values_only=True),
            aliases=HISTORY_HEADER_ALIASES,
            required_fields=("nome", "competencia"),
        )
        if not mapping or "valor" in mapping.columns or "historico_nome" in mapping.columns:
            continue

        header_row = next(worksheet.iter_rows(min_row=mapping.header_row, max_row=mapping.header_row, values_only=True))
        header = list(header_row)
        competencia_index = mapping.columns["competencia"]
        value_columns = [
            index
            for index, value in enumerate(header)
            if index > competencia_index and index not in mapping.columns.values() and to_optional_str(value)
        ]
        if not value_columns:
            continue

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=mapping.header_row + 1, values_only=True),
            start=mapping.header_row + 1,
        ):
            nome = to_optional_str(cell_value(row, mapping.columns.get("nome")))
            competencia = normalize_competencia(cell_value(row, mapping.columns.get("competencia")))
            matricula = normalize_registration(cell_value(row, mapping.columns.get("matricula")))
            if not any((nome, competencia, matricula)):
                continue
            if not nome or not competencia:
                invalid_rows.append(f"{worksheet.title}:{row_index}")
                continue

            for column_index in value_columns:
                raw_valor = cell_value(row, column_index)
                if not has_meaningful_history_value(raw_valor):
                    continue

                historico_nome = clean_history_label(header[column_index])
                try:
                    valor = parse_decimal(raw_valor)
                except ValueError:
                    invalid_rows.append(f"{worksheet.title}:{row_index}")
                    continue

                if matricula:
                    add_history_value(
                        history_index.by_matricula,
                        matricula,
                        historico_nome,
                        competencia,
                        valor,
                    )
                add_history_value(
                    history_index.by_nome,
                    normalize_name_key(nome),
                    historico_nome,
                    competencia,
                    valor,
                )

    return invalid_rows


def resolve_historicos(
    cpf: str,
    matricula: str | None,
    nome: str,
    historico_nome: str | None,
    history_index: HistoryIndex,
) -> list[HistoricalSeries]:
    series_by_name = history_index.by_cpf.get(cpf, {})
    if not series_by_name and matricula:
        series_by_name = history_index.by_matricula.get(matricula, {})
    if not series_by_name and nome:
        series_by_name = history_index.by_nome.get(normalize_name_key(nome), {})

    if historico_nome:
        match = series_by_name.get(normalize_name_key(historico_nome))
        if match:
            return [match]
        return [HistoricalSeries(nome=historico_nome, valores=[])]

    return list(series_by_name.values())


def history_sources(workbook: Workbook, history_workbook: Workbook | None) -> list[Workbook]:
    if history_workbook is None or history_workbook is workbook:
        return [workbook]
    return [workbook, history_workbook]


def sort_history_values(history_index: HistoryIndex) -> None:
    for groups in (history_index.by_cpf, history_index.by_matricula, history_index.by_nome):
        for series_by_name in groups.values():
            for series in series_by_name.values():
                series.valores.sort(key=lambda item: competencia_sort_key(item.competencia))


def add_history_value(
    index: dict[str, dict[str, HistoricalSeries]],
    owner_key: str,
    historico_nome: str,
    competencia: str,
    valor: Decimal,
) -> None:
    series_by_name = index.setdefault(owner_key, {})
    normalized_name = normalize_name_key(historico_nome)
    series = series_by_name.setdefault(normalized_name, HistoricalSeries(nome=historico_nome.strip(), valores=[]))
    series.valores.append(HistoricalValue(competencia=competencia, valor=valor))


def clean_history_label(value: object) -> str:
    return " ".join(str(value or "").split())


def has_meaningful_history_value(value: object) -> bool:
    text = str(value or "").strip()
    return bool(text) and text not in {"-", "--", "—"}


def competencia_sort_key(competencia: str) -> tuple[int, int, str]:
    try:
        parsed = datetime.strptime(competencia, "%m/%Y")
        return (parsed.year, parsed.month, competencia)
    except ValueError:
        return (9999, 12, competencia)


def cell_value(row: tuple[object, ...], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def to_optional_str(value: object) -> str | None:
    text = str(value).strip() if value is not None else ""
    return text or None
