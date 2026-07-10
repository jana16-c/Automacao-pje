from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
import re

from openpyxl.workbook.workbook import Workbook

from pje_automation.domain.execution import ExecutionMode
from pje_automation.domain.models import HistoricalSeries, HistoricalValue, Record, RecordSource, WorkbookPreview
from pje_automation.excel.normalization import (
    normalize_competencia,
    normalize_cpf,
    normalize_date,
    normalize_header,
    normalize_name_key,
    normalize_process_digits,
    normalize_registration,
    parse_decimal,
)


MAIN_HEADER_ALIASES = {
    "matricula": {"matricula", "registro", "codigo", "codigo empregado"},
    "nome": {"nome", "reclamante", "empregado", "nome reclamante"},
    "cpf": {"cpf", "cpf reclamante", "documento", "documento fiscal"},
    "data_admissao": {"admissao", "data admissao", "dt admissao"},
    "data_demissao": {"demissao", "data demissao", "dt demissao", "data final"},
    "data_calculo": {"data calculo", "dt calculo", "data de calculo", "data termino calculo"},
    "processo": {"processo", "numero processo", "numero do processo"},
    "historico_nome": {"historico nome", "historico salarial", "nome historico", "verba historico"},
}

HISTORY_HEADER_ALIASES = {
    "cpf": MAIN_HEADER_ALIASES["cpf"],
    "matricula": MAIN_HEADER_ALIASES["matricula"],
    "historico_nome": MAIN_HEADER_ALIASES["historico_nome"],
    "nome": MAIN_HEADER_ALIASES["nome"] | {"funcionario"},
    "competencia": {"competencia", "mes ano", "mes/ano", "referencia", "periodo"},
    "valor": {"valor", "valor historico", "valor base", "base informada"},
}


@dataclass(slots=True)
class SheetMapping:
    header_row: int
    columns: dict[str, int]


@dataclass(slots=True)
class HistoryIndex:
    by_cpf: dict[str, dict[str, HistoricalSeries]]
    by_matricula: dict[str, dict[str, HistoricalSeries]]
    by_nome: dict[str, dict[str, HistoricalSeries]]


@dataclass(slots=True)
class ValueColumnProfile:
    numeric_count: int = 0
    one_decimal_count: int = 0
    two_plus_decimal_count: int = 0


@dataclass(slots=True)
class PendingRecord:
    record_id: str
    nome: str
    cpf: str
    matricula: str | None
    data_admissao: str | None
    data_demissao: str | None
    data_calculo: str | None
    processo: str | None
    historico_nome: str | None
    source: RecordSource


@dataclass(slots=True)
class HistoryLookupTargets:
    matriculas: set[str]
    nomes: set[str]


def build_preview(
    workbook: Workbook,
    history_workbook: Workbook | None = None,
    limit: int | None = 20,
    execution_mode: ExecutionMode = ExecutionMode.NOVO_CALCULO,
    fixed_process: str | None = None,
) -> WorkbookPreview:
    valid_records: list[Record] = []
    invalid_rows: list[str] = []
    ambiguous_rows: list[str] = []
    pending_records, control_invalid_rows, used_control_sheets = collect_control_records(
        workbook,
        limit=limit,
        execution_mode=execution_mode,
        fixed_process=fixed_process,
    )
    invalid_rows.extend(control_invalid_rows)

    targets = build_history_lookup_targets(pending_records)
    history_index, history_invalid_rows, used_history_sheets = collect_history_series(
        workbook,
        history_workbook=history_workbook,
        targets=targets,
    )
    invalid_rows.extend(history_invalid_rows)

    for pending in pending_records:
        valid_records.append(
            Record(
                record_id=pending.record_id,
                nome=pending.nome,
                cpf=pending.cpf,
                data_admissao=pending.data_admissao,
                data_demissao=pending.data_demissao,
                data_calculo=pending.data_calculo,
                processo=pending.processo,
                historicos=resolve_historicos(
                    cpf=pending.cpf,
                    matricula=pending.matricula,
                    nome=pending.nome,
                    historico_nome=pending.historico_nome,
                    history_index=history_index,
                ),
                source=pending.source,
            )
        )

    if not valid_records:
        ambiguous_rows.append(empty_preview_message(execution_mode))

    if history_workbook is None or history_workbook is workbook:
        ignored_control_sheets = sorted(sheet for sheet in workbook.sheetnames if sheet not in (used_control_sheets | used_history_sheets))
        ignored_history_sheets: list[str] = []
    else:
        ignored_control_sheets = sorted(sheet for sheet in workbook.sheetnames if sheet not in used_control_sheets)
        ignored_history_sheets = sorted(sheet for sheet in history_workbook.sheetnames if sheet not in used_history_sheets)

    return WorkbookPreview(
        valid_records=valid_records,
        invalid_rows=invalid_rows,
        ambiguous_rows=ambiguous_rows,
        sheet_names=workbook.sheetnames,
        ignored_control_sheets=ignored_control_sheets,
        ignored_history_sheets=ignored_history_sheets,
    )


def collect_control_records(
    workbook: Workbook,
    limit: int | None = 20,
    execution_mode: ExecutionMode = ExecutionMode.NOVO_CALCULO,
    fixed_process: str | None = None,
) -> tuple[list[PendingRecord], list[str], set[str]]:
    pending_records: list[PendingRecord] = []
    invalid_rows: list[str] = []
    used_control_sheets: set[str] = set()

    for worksheet in workbook.worksheets:
        mapping = detect_control_mapping(worksheet, execution_mode=execution_mode)
        if not mapping:
            continue
        used_control_sheets.add(worksheet.title)

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=max(mapping.header_row + 1, 1), values_only=True),
            start=max(mapping.header_row + 1, 1),
        ):
            nome = to_optional_str(cell_value(row, mapping.columns.get("nome")))
            if not nome:
                continue

            cpf = normalize_cpf(cell_value(row, mapping.columns.get("cpf")))
            matricula = normalize_registration(cell_value(row, mapping.columns.get("matricula")))
            processo = to_optional_str(cell_value(row, mapping.columns.get("processo"))) or to_optional_str(fixed_process)
            data_demissao = normalize_date(cell_value(row, mapping.columns.get("data_demissao")))
            data_calculo = normalize_date(cell_value(row, mapping.columns.get("data_calculo")))
            if not is_control_row_valid(
                execution_mode,
                cpf=cpf,
                processo=processo,
                data_demissao=data_demissao,
                data_calculo=data_calculo,
            ):
                invalid_rows.append(f"{worksheet.title}:{row_index}")
                continue

            record_id = build_record_id(
                execution_mode,
                cpf=cpf,
                nome=nome,
                matricula=matricula,
                processo=processo,
                worksheet_title=worksheet.title,
                row_index=row_index,
            )
            pending_records.append(
                PendingRecord(
                    record_id=record_id,
                    nome=nome,
                    cpf=cpf,
                    matricula=matricula,
                    data_admissao=normalize_date(cell_value(row, mapping.columns.get("data_admissao"))),
                    data_demissao=data_demissao,
                    data_calculo=data_calculo,
                    processo=processo,
                    historico_nome=to_optional_str(cell_value(row, mapping.columns.get("historico_nome"))),
                    source=RecordSource(sheet=worksheet.title, row=row_index),
                )
            )
            if limit is not None and len(pending_records) >= limit:
                return pending_records, invalid_rows, used_control_sheets
    return pending_records, invalid_rows, used_control_sheets


def detect_control_mapping(worksheet, execution_mode: ExecutionMode = ExecutionMode.NOVO_CALCULO) -> SheetMapping | None:
    mapping = detect_mapping(
        worksheet.iter_rows(values_only=True),
        aliases=MAIN_HEADER_ALIASES,
        required_fields=control_required_fields(execution_mode),
    )
    if mapping:
        return mapping
    return infer_headerless_control_mapping(worksheet, execution_mode)


def infer_headerless_control_mapping(worksheet, execution_mode: ExecutionMode) -> SheetMapping | None:
    for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
        if not row:
            continue
        matricula = normalize_registration(cell_value(row, 0))
        nome = to_optional_str(cell_value(row, 1))
        if not matricula or not nome:
            continue

        columns = {"matricula": 0, "nome": 1}
        if cpf_column := headerless_cpf_column(row):
            columns["cpf"] = cpf_column
        if process_column := headerless_process_column(row):
            columns["processo"] = process_column

        date_columns = headerless_date_columns(row, skip_columns=set(columns.values()))
        if execution_mode == ExecutionMode.NOVO_CALCULO:
            if "cpf" not in columns:
                continue
            if len(date_columns) >= 2:
                columns["data_admissao"] = date_columns[0]
                columns["data_demissao"] = date_columns[1]
            elif len(date_columns) == 1:
                columns["data_demissao"] = date_columns[0]
            if "data_demissao" not in columns:
                continue
            return SheetMapping(header_row=0, columns=columns)

        if date_columns:
            columns["data_demissao"] = date_columns[0]
        if execution_mode == ExecutionMode.CORRIGIR_DATAS_E_HISTORICO and len(date_columns) >= 2:
            columns["data_calculo"] = date_columns[1]
        return SheetMapping(header_row=0, columns=columns)
    return None


def control_required_fields(execution_mode: ExecutionMode) -> tuple[str, ...]:
    if execution_mode == ExecutionMode.NOVO_CALCULO:
        return ("nome", "cpf", "data_demissao")
    return ("nome", "processo")


def is_control_row_valid(
    execution_mode: ExecutionMode,
    *,
    cpf: str,
    processo: str | None,
    data_demissao: str | None,
    data_calculo: str | None,
) -> bool:
    if execution_mode == ExecutionMode.NOVO_CALCULO:
        return bool(cpf and data_demissao)
    if execution_mode == ExecutionMode.CORRIGIR_DATAS_E_HISTORICO:
        return bool(data_demissao or data_calculo)
    return True


def build_record_id(
    execution_mode: ExecutionMode,
    *,
    cpf: str,
    nome: str,
    matricula: str | None,
    processo: str | None,
    worksheet_title: str,
    row_index: int,
) -> str:
    if execution_mode == ExecutionMode.NOVO_CALCULO and cpf:
        return cpf
    process_digits = normalize_process_digits(processo)
    if process_digits:
        return f"{process_digits}-{normalize_name_key(nome)}"
    if matricula:
        return f"{matricula}-{normalize_name_key(nome)}"
    return f"{worksheet_title}-{row_index}"


def empty_preview_message(execution_mode: ExecutionMode) -> str:
    if execution_mode == ExecutionMode.CORRIGIR_HISTORICO:
        return "Nenhuma linha elegivel foi encontrada com nome ou matricula para corrigir o historico."
    if execution_mode == ExecutionMode.CORRIGIR_DATAS_E_HISTORICO:
        return "Nenhuma linha elegivel foi encontrada com nome ou matricula e ao menos uma data para corrigir."
    return "Nenhuma linha elegivel foi encontrada com nome, CPF e data de demissao."


def headerless_cpf_column(row: tuple[object, ...]) -> int | None:
    for index in range(2, min(len(row), 6)):
        value = cell_value(row, index)
        if looks_like_headerless_cpf(value):
            return index
    return None


def headerless_process_column(row: tuple[object, ...]) -> int | None:
    for index in range(2, min(len(row), 6)):
        if len(normalize_process_digits(cell_value(row, index))) >= 16:
            return index
    return None


def headerless_date_columns(row: tuple[object, ...], skip_columns: set[int]) -> list[int]:
    columns: list[int] = []
    for index in range(2, min(len(row), 7)):
        if index in skip_columns:
            continue
        if looks_like_control_date(cell_value(row, index)):
            columns.append(index)
    return columns


def looks_like_headerless_cpf(value: object) -> bool:
    if value in (None, ""):
        return False
    if looks_like_control_date(value):
        return False
    text = str(value).strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) != 11:
        return False
    return any(marker in text for marker in (".", "-", "/")) or text.isdigit()


def build_history_lookup_targets(records: list[PendingRecord]) -> HistoryLookupTargets:
    return HistoryLookupTargets(
        matriculas={record.matricula for record in records if record.matricula},
        nomes={normalize_name_key(record.nome) for record in records if record.nome},
    )


def detect_mapping(
    rows: Iterable[tuple[object, ...]],
    aliases: dict[str, set[str]],
    required_fields: tuple[str, ...],
) -> SheetMapping | None:
    for row_index, row in enumerate(rows, start=1):
        columns: dict[str, int] = {}
        for index, value in enumerate(row):
            header = normalize_header(value)
            if not header:
                continue
            for field, accepted_headers in aliases.items():
                if header in accepted_headers:
                    columns[field] = index
        if all(field in columns for field in required_fields):
            return SheetMapping(header_row=row_index, columns=columns)
        if row_index >= 15:
            break
    return None


def collect_history_series(
    workbook: Workbook,
    history_workbook: Workbook | None = None,
    targets: HistoryLookupTargets | None = None,
) -> tuple[HistoryIndex, list[str], set[str]]:
    history_index = HistoryIndex(by_cpf={}, by_matricula={}, by_nome={})
    invalid_rows: list[str] = []
    used_sheets: set[str] = set()

    for source_workbook in history_sources(workbook, history_workbook):
        structured_invalid_rows, structured_used_sheets = collect_structured_history_series(source_workbook, history_index, targets=targets)
        invalid_rows.extend(structured_invalid_rows)
        used_sheets.update(structured_used_sheets)
        companion_invalid_rows, companion_used_sheets = collect_companion_history_series(source_workbook, history_index, targets=targets)
        invalid_rows.extend(companion_invalid_rows)
        used_sheets.update(companion_used_sheets)

    sort_history_values(history_index)
    return history_index, invalid_rows, used_sheets


def collect_structured_history_series(
    workbook: Workbook,
    history_index: HistoryIndex,
    targets: HistoryLookupTargets | None = None,
) -> tuple[list[str], set[str]]:
    invalid_rows: list[str] = []
    used_sheets: set[str] = set()
    for worksheet in workbook.worksheets:
        if should_skip_history_sheet_by_title(workbook, worksheet.title, targets):
            continue
        mapping = detect_mapping(
            worksheet.iter_rows(values_only=True),
            aliases=HISTORY_HEADER_ALIASES,
            required_fields=("historico_nome", "competencia", "valor"),
        )
        if not mapping:
            continue
        used_sheets.add(worksheet.title)

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=mapping.header_row + 1, values_only=True),
            start=mapping.header_row + 1,
        ):
            cpf = normalize_cpf(cell_value(row, mapping.columns.get("cpf")))
            matricula = normalize_registration(cell_value(row, mapping.columns.get("matricula")))
            nome = to_optional_str(cell_value(row, mapping.columns.get("nome")))
            historico_nome = to_optional_str(cell_value(row, mapping.columns.get("historico_nome")))
            competencia = normalize_competencia(cell_value(row, mapping.columns.get("competencia")))
            raw_valor = cell_value(row, mapping.columns.get("valor"))

            if not any((cpf, matricula, nome, historico_nome, competencia, raw_valor)):
                continue
            if not historico_nome or not competencia or not any((matricula, nome, cpf)):
                invalid_rows.append(f"{worksheet.title}:{row_index}")
                continue
            if targets and not matches_history_target(targets, matricula=matricula, nome=nome):
                continue
            if not has_meaningful_history_value(raw_valor):
                continue

            try:
                valor = parse_decimal(raw_valor)
            except ValueError:
                invalid_rows.append(f"{worksheet.title}:{row_index}")
                continue

            if matricula:
                add_history_value(history_index.by_matricula, matricula, historico_nome, competencia, valor)
            if nome:
                add_history_value(history_index.by_nome, normalize_name_key(nome), historico_nome, competencia, valor)
            if cpf:
                add_history_value(history_index.by_cpf, cpf, historico_nome, competencia, valor)

    return invalid_rows, used_sheets


def collect_companion_history_series(
    workbook: Workbook,
    history_index: HistoryIndex,
    targets: HistoryLookupTargets | None = None,
) -> tuple[list[str], set[str]]:
    invalid_rows: list[str] = []
    used_sheets: set[str] = set()
    for worksheet in workbook.worksheets:
        if should_skip_history_sheet_by_title(workbook, worksheet.title, targets):
            continue
        mapping = detect_mapping(
            worksheet.iter_rows(values_only=True),
            aliases=HISTORY_HEADER_ALIASES,
            required_fields=("competencia",),
        )
        if not mapping or "valor" in mapping.columns or "historico_nome" in mapping.columns:
            continue
        used_sheets.add(worksheet.title)

        header_row = next(worksheet.iter_rows(min_row=mapping.header_row, max_row=mapping.header_row, values_only=True))
        header = list(header_row)
        competencia_index = mapping.columns["competencia"]
        value_columns = [
            index
            for index, value in enumerate(header)
            if index > competencia_index and index not in mapping.columns.values() and to_optional_str(value)
        ]
        if not value_columns:
            continue

        profiles = build_value_column_profiles(worksheet, mapping.header_row, value_columns)
        sheet_matricula = sheet_registration_hint(worksheet, mapping)
        sheet_nome = sheet_name_hint(worksheet, mapping)
        if targets and not matches_history_target(targets, matricula=sheet_matricula, nome=sheet_nome):
            continue

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=mapping.header_row + 1, values_only=False),
            start=mapping.header_row + 1,
        ):
            nome = to_optional_str(cell_value(row, mapping.columns.get("nome"))) or sheet_nome
            competencia = normalize_competencia(cell_value(row, mapping.columns.get("competencia")))
            matricula = normalize_registration(cell_value(row, mapping.columns.get("matricula"))) or sheet_matricula
            if not any((nome, competencia, matricula)):
                continue
            if not competencia or not (nome or matricula):
                invalid_rows.append(f"{worksheet.title}:{row_index}")
                continue

            for column_index in value_columns:
                raw_valor = cell_value(row, column_index)
                if not has_meaningful_history_value(raw_valor):
                    continue

                historico_nome = clean_history_label(header[column_index])
                try:
                    valor = parse_history_value(row[column_index], profiles.get(column_index))
                except ValueError:
                    invalid_rows.append(f"{worksheet.title}:{row_index}")
                    continue

                if matricula:
                    add_history_value(
                        history_index.by_matricula,
                        matricula,
                        historico_nome,
                        competencia,
                        valor,
                    )
                if nome:
                    add_history_value(
                        history_index.by_nome,
                        normalize_name_key(nome),
                        historico_nome,
                        competencia,
                        valor,
                    )

    return invalid_rows, used_sheets


def resolve_historicos(
    cpf: str,
    matricula: str | None,
    nome: str,
    historico_nome: str | None,
    history_index: HistoryIndex,
) -> list[HistoricalSeries]:
    series_by_name: dict[str, HistoricalSeries] = {}
    if matricula:
        series_by_name = history_index.by_matricula.get(matricula, {})
    if not series_by_name and nome:
        series_by_name = history_index.by_nome.get(normalize_name_key(nome), {})

    if historico_nome:
        match = series_by_name.get(normalize_name_key(historico_nome))
        if match:
            return [match]
        return [HistoricalSeries(nome=historico_nome, valores=[])]

    return list(series_by_name.values())


def history_sources(workbook: Workbook, history_workbook: Workbook | None) -> list[Workbook]:
    if history_workbook is None or history_workbook is workbook:
        return [workbook]
    return [workbook, history_workbook]


def sheet_registration_hint(worksheet, mapping: SheetMapping) -> str | None:
    row_after_header = mapping.header_row + 1
    if "matricula" in mapping.columns:
        value = worksheet.cell(row=row_after_header, column=mapping.columns["matricula"] + 1).value
        matricula = normalize_registration(value)
        if matricula:
            return matricula
    return normalize_registration(worksheet.cell(row=2, column=1).value)


def sheet_name_hint(worksheet, mapping: SheetMapping) -> str | None:
    row_after_header = mapping.header_row + 1
    if "nome" in mapping.columns:
        nome = to_optional_str(worksheet.cell(row=row_after_header, column=mapping.columns["nome"] + 1).value)
        if nome:
            return nome
    title = " ".join(str(worksheet.title or "").split())
    if not title or normalize_name_key(title).startswith("PLANILHA"):
        return None
    return title


def matches_history_target(targets: HistoryLookupTargets, matricula: str | None, nome: str | None) -> bool:
    if matricula and matricula in targets.matriculas:
        return True
    if nome and normalize_name_key(nome) in targets.nomes:
        return True
    return False


def should_skip_history_sheet_by_title(workbook: Workbook, title: str, targets: HistoryLookupTargets | None) -> bool:
    if targets is None or len(workbook.sheetnames) < 100:
        return False
    normalized_title = normalize_name_key(title or "")
    if not normalized_title:
        return False
    if normalized_title.startswith("PLANILHA") or normalized_title.startswith("HISTORICO") or normalized_title.startswith("RESUMO"):
        return False
    return normalized_title not in targets.nomes


def looks_like_control_date(value: object) -> bool:
    if value in (None, ""):
        return False
    if isinstance(value, datetime):
        return True
    if isinstance(value, (int, float)):
        return True
    text = str(value).strip()
    return bool(re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text))


def sort_history_values(history_index: HistoryIndex) -> None:
    for groups in (history_index.by_cpf, history_index.by_matricula, history_index.by_nome):
        for series_by_name in groups.values():
            for series in series_by_name.values():
                series.valores.sort(key=lambda item: competencia_sort_key(item.competencia))


def add_history_value(
    index: dict[str, dict[str, HistoricalSeries]],
    owner_key: str,
    historico_nome: str,
    competencia: str,
    valor: Decimal,
) -> None:
    series_by_name = index.setdefault(owner_key, {})
    normalized_name = normalize_name_key(historico_nome)
    series = series_by_name.setdefault(normalized_name, HistoricalSeries(nome=historico_nome.strip(), valores=[]))
    series.valores.append(HistoricalValue(competencia=competencia, valor=valor))


def build_value_column_profiles(worksheet, header_row: int, value_columns: list[int]) -> dict[int, ValueColumnProfile]:
    profiles = {column_index: ValueColumnProfile() for column_index in value_columns}
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=False):
        for column_index in value_columns:
            raw_value = cell_value(row, column_index)
            if not has_meaningful_history_value(raw_value):
                continue
            decimal_places = decimal_places_count(raw_value)
            if decimal_places is None:
                continue
            profile = profiles[column_index]
            profile.numeric_count += 1
            if decimal_places == 1:
                profile.one_decimal_count += 1
            elif decimal_places >= 2:
                profile.two_plus_decimal_count += 1
    return profiles


def parse_history_value(raw_value: object, profile: ValueColumnProfile | None) -> Decimal:
    value = parse_decimal(raw_cell_value(raw_value))
    if profile and should_shift_sparse_single_decimal(raw_value, profile):
        return (value / Decimal("10")).quantize(Decimal("0.01"))
    return value


def should_shift_sparse_single_decimal(raw_value: object, profile: ValueColumnProfile) -> bool:
    formatted_places = formatted_decimal_places_count(raw_value)
    if formatted_places is not None and formatted_places >= 2:
        return False
    if profile.one_decimal_count <= 0 or profile.two_plus_decimal_count <= 0:
        return False
    if profile.one_decimal_count > max(3, profile.numeric_count // 5):
        return False
    if decimal_places_count(raw_value) != 1:
        return False
    try:
        return Decimal(str(raw_cell_value(raw_value))).copy_abs() < Decimal("10")
    except Exception:
        return False


def decimal_places_count(value: object) -> int | None:
    value = raw_cell_value(value)
    if isinstance(value, int):
        return 0
    if isinstance(value, float):
        text = format(value, ".15g")
    elif isinstance(value, Decimal):
        text = format(value.normalize(), "f")
    else:
        text = str(value).strip()
    if not text or "." not in text:
        return 0 if text else None
    return len(text.split(".", 1)[1].rstrip("0"))


def clean_history_label(value: object) -> str:
    return " ".join(str(value or "").split())


def has_meaningful_history_value(value: object) -> bool:
    text = str(value or "").strip()
    return bool(text) and text not in {"-", "--", "—"}


def competencia_sort_key(competencia: str) -> tuple[int, int, str]:
    try:
        parsed = datetime.strptime(competencia, "%m/%Y")
        return (parsed.year, parsed.month, competencia)
    except ValueError:
        return (9999, 12, competencia)


def cell_value(row: tuple[object, ...], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return raw_cell_value(row[index])


def raw_cell_value(value: object) -> object:
    return getattr(value, "value", value)


def formatted_decimal_places_count(value: object) -> int | None:
    number_format = getattr(value, "number_format", None)
    if not number_format or number_format == "General":
        return None
    section = str(number_format).split(";", 1)[0]
    if "." in section:
        decimal_part = section.rsplit(".", 1)[1]
    elif "," in section and any(marker in section.rsplit(",", 1)[1] for marker in ("0", "#")):
        decimal_part = section.rsplit(",", 1)[1]
    else:
        return 0
    count = 0
    for char in decimal_part:
        if char not in {"0", "#"}:
            break
        count += 1
    return count


def to_optional_str(value: object) -> str | None:
    text = str(value).strip() if value is not None else ""
    return text or None
