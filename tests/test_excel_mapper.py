from openpyxl import Workbook

from pje_automation.domain.execution import ExecutionMode
from pje_automation.excel.mapper import build_preview


def test_build_preview_attaches_history_rows_by_registration() -> None:
    workbook = Workbook()
    dados = workbook.active
    dados.title = "Dados"
    dados.append(["Matricula", "Nome", "CPF", "Data Demissao"])
    dados.append(["1001", "Maria Souza", "12345678901", "30/06/2026"])

    historico = workbook.create_sheet("Historico_Salarial")
    historico.append(["Matricula", "Historico Nome", "Competencia", "Valor"])
    historico.append(["1001", "DIFERENCA", "07/2012", "10,50"])
    historico.append(["1001", "DIFERENCA", "08/2012", "11,75"])

    preview = build_preview(workbook)

    assert len(preview.valid_records) == 1
    record = preview.valid_records[0]
    assert len(record.historicos) == 1
    assert record.historicos[0].nome == "DIFERENCA"
    assert [item.competencia for item in record.historicos[0].valores] == ["07/2012", "08/2012"]


def test_build_preview_filters_history_by_requested_name() -> None:
    workbook = Workbook()
    dados = workbook.active
    dados.title = "Dados"
    dados.append(["Matricula", "Nome", "CPF", "Data Demissao", "Historico Nome"])
    dados.append(["1001", "Maria Souza", "12345678901", "30/06/2026", "BASE INFORMADA"])

    historico = workbook.create_sheet("Historico_Salarial")
    historico.append(["Matricula", "Historico Nome", "Competencia", "Valor"])
    historico.append(["1001", "BASE INFORMADA", "07/2012", "10,50"])
    historico.append(["1001", "OUTRA SERIE", "07/2012", "99,99"])

    preview = build_preview(workbook)

    record = preview.valid_records[0]
    assert [serie.nome for serie in record.historicos] == ["BASE INFORMADA"]


def test_build_preview_returns_placeholder_when_requested_history_is_missing() -> None:
    workbook = Workbook()
    dados = workbook.active
    dados.title = "Dados"
    dados.append(["Matricula", "Nome", "CPF", "Data Demissao", "Historico Nome"])
    dados.append(["1001", "Maria Souza", "12345678901", "30/06/2026", "BASE INFORMADA"])

    preview = build_preview(workbook)

    record = preview.valid_records[0]
    assert len(record.historicos) == 1
    assert record.historicos[0].nome == "BASE INFORMADA"
    assert record.historicos[0].valores == []


def test_build_preview_supports_companion_history_workbook_and_name_fallback() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Admissao", "Demissao"])
    controle.append(["1006133", "ATHOS HENRIQUE MENDES SILVA", "081.367.163-90", "06/08/2015", "24/03/2016"])

    historico = Workbook()
    aba = historico.active
    aba.title = "ATHOS HENRIQUE MENDES SILVA"
    aba.append(["Matricula", "Funcionário", "Dt.Admissão", "Dt.Demissão", "Período", "dif ad.not+ \nReflexos"])
    aba.append(["1007592", "ATHOS HENRIQUE MENDES SILVA", "06/08/2015", "24/03/16", "06/2015", "18,54"])
    aba.append(["1007592", "ATHOS HENRIQUE MENDES SILVA", "06/08/2015", "24/03/16", "07/2015", "37,18"])
    aba.append(["1007592", "ATHOS HENRIQUE MENDES SILVA", "06/08/2015", "24/03/16", "08/2015", "-"])

    preview = build_preview(cadastro, history_workbook=historico)

    assert len(preview.valid_records) == 1
    record = preview.valid_records[0]
    assert record.nome == "ATHOS HENRIQUE MENDES SILVA"
    assert record.cpf == "08136716390"
    assert len(record.historicos) == 1
    assert record.historicos[0].nome == "dif ad.not+ Reflexos"
    assert [item.competencia for item in record.historicos[0].valores] == ["06/2015", "07/2015"]


def test_build_preview_prioritizes_registration_before_name() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Demissao"])
    controle.append(["1001", "MARIA SOUZA", "12345678901", "30/06/2026"])

    historico = Workbook()
    aba = historico.active
    aba.title = "Historico"
    aba.append(["Matricula", "Nome", "Historico Nome", "Competencia", "Valor"])
    aba.append(["9999", "MARIA SOUZA", "BASE ERRADA", "01/2020", "99,99"])
    aba.append(["1001", "OUTRA PESSOA", "BASE CERTA", "01/2020", "10,00"])

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    record = preview.valid_records[0]
    assert [serie.nome for serie in record.historicos] == ["BASE CERTA"]


def test_build_preview_matches_companion_history_across_multiple_employee_sheets_by_a2_registration() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Demissao"])
    controle.append(["1001", "ANA SILVA", "11111111111", "01/02/2020"])
    controle.append(["1002", "BRUNO LIMA", "22222222222", "01/02/2020"])

    historico = Workbook()
    ana = historico.active
    ana.title = "ANA SILVA"
    ana["A1"] = "Matricula"
    ana["A2"] = "1001"
    ana.append([])
    ana.append(["Periodo", "Base"])
    ana.append(["01/2020", "10,00"])
    ana.append(["02/2020", "20,00"])

    bruno = historico.create_sheet("BRUNO LIMA")
    bruno["A1"] = "Matricula"
    bruno["A2"] = "1002"
    bruno.append([])
    bruno.append(["Periodo", "Base"])
    bruno.append(["01/2020", "30,00"])

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    records_by_name = {record.nome: record for record in preview.valid_records}
    ana_values = [str(item.valor) for item in records_by_name["ANA SILVA"].historicos[0].valores]
    bruno_values = [str(item.valor) for item in records_by_name["BRUNO LIMA"].historicos[0].valores]

    assert ana_values == ["10.00", "20.00"]
    assert bruno_values == ["30.00"]


def test_build_preview_matches_companion_history_by_first_three_name_tokens_when_sheet_title_is_truncated() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Demissao"])
    controle.append(["1001", "JOAO PEDRO ALVES SANTOS", "11111111111", "01/02/2020"])

    historico = Workbook()
    aba = historico.active
    aba.title = "JOAO PEDRO ALVES SANTO"
    aba.append([])
    aba.append([])
    aba.append(["Periodo", "Base"])
    aba.append(["01/2020", "10,00"])
    aba.append(["02/2020", "20,00"])

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    record = preview.valid_records[0]
    valores = [str(item.valor) for item in record.historicos[0].valores]

    assert valores == ["10.00", "20.00"]


def test_build_preview_matches_companion_history_ignoring_particles_and_last_token_truncation() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Demissao"])
    controle.append(["1001", "MARIA DE FATIMA SOUZA", "11111111111", "01/02/2020"])

    historico = Workbook()
    aba = historico.active
    aba.title = "MARIA FATIMA SOUZ"
    aba.append([])
    aba.append([])
    aba.append(["Periodo", "Base"])
    aba.append(["01/2020", "10,00"])
    aba.append(["02/2020", "20,00"])

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    record = preview.valid_records[0]
    valores = [str(item.valor) for item in record.historicos[0].valores]

    assert valores == ["10.00", "20.00"]


def test_build_preview_keeps_large_history_workbook_sheet_when_a2_registration_matches() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Demissao"])
    controle.append(["1001", "MARIA DE FATIMA SOUZA", "11111111111", "01/02/2020"])

    historico = Workbook()
    aba = historico.active
    aba.title = "FUNCIONARIO 1001"
    aba["A1"] = "Matricula"
    aba["A2"] = "1001"
    aba.append([])
    aba.append(["Periodo", "Base"])
    aba.append(["01/2020", "10,00"])
    aba.append(["02/2020", "20,00"])

    for index in range(2, 103):
        filler = historico.create_sheet(f"Planilha {index}")
        filler.append(["Observacao"])
        filler.append([f"aba {index}"])

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    record = preview.valid_records[0]
    valores = [str(item.valor) for item in record.historicos[0].valores]

    assert valores == ["10.00", "20.00"]


def test_build_preview_lists_ignored_history_sheets() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Demissao"])
    controle.append(["1001", "ANA SILVA", "11111111111", "01/02/2020"])

    historico = Workbook()
    ana = historico.active
    ana.title = "ANA SILVA"
    ana.append(["Matricula", "Funcionario", "Periodo", "Base"])
    ana.append(["1001", "ANA SILVA", "01/2020", "10,00"])
    historico.create_sheet("Resumo")

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    assert preview.ignored_history_sheets == ["Resumo"]


def test_build_preview_without_limit_keeps_rows_after_the_default_sample() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Admissao", "Demissao"])
    for idx in range(1, 26):
        controle.append([f"{1000 + idx}", f"Pessoa {idx}", f"{idx:011d}", "01/01/2020", "02/01/2020"])

    historico = Workbook()
    aba = historico.active
    aba.title = "Pessoa 25"
    aba.append(["Matricula", "Funcionário", "Dt.Admissão", "Dt.Demissão", "Período", "Base"])
    aba.append(["1025", "Pessoa 25", "01/01/2020", "02/01/2020", "01/2020", "10,00"])

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    assert len(preview.valid_records) == 25
    assert preview.valid_records[-1].nome == "Pessoa 25"
    assert preview.valid_records[-1].historicos[0].valores[0].competencia == "01/2020"


def test_build_preview_normalizes_sparse_single_decimal_history_values() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Admissao", "Demissao"])
    controle.append(["1004664", "FILIPE", "279.892.794-36", "02/09/2004", "11/10/2019"])

    historico = Workbook()
    aba = historico.active
    aba.title = "FILIPE"
    aba.append(["Matricula", "Funcionário", "Dt.Admissão", "Dt.Demissão", "Período", "Base"])
    aba.append(["1004664", "FILIPE", "02/09/2004", "11/10/2019", "09/2014", 0.24])
    aba.append(["1004664", "FILIPE", "02/09/2004", "11/10/2019", "10/2014", 0.1])
    aba.append(["1004664", "FILIPE", "02/09/2004", "11/10/2019", "11/2014", 0.18])
    aba.append(["1004664", "FILIPE", "02/09/2004", "11/10/2019", "01/2015", 1.3])
    aba.append(["1004664", "FILIPE", "02/09/2004", "11/10/2019", "02/2015", 0.22])
    aba.append(["1004664", "FILIPE", "02/09/2004", "11/10/2019", "12/2017", 0.1])
    aba.append(["1004664", "FILIPE", "02/09/2004", "11/10/2019", "01/2018", 0.11])

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    valores = {item.competencia: str(item.valor) for serie in preview.valid_records[0].historicos for item in serie.valores}
    assert valores["10/2014"] == "0.01"
    assert valores["01/2015"] == "0.13"
    assert valores["12/2017"] == "0.01"


def test_build_preview_preserves_two_decimal_formatted_history_values() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Admissao", "Demissao"])
    controle.append(["1001", "Maria", "12345678901", "01/01/2020", "02/01/2020"])

    historico = Workbook()
    aba = historico.active
    aba.title = "Maria"
    aba.append(["Matricula", "Funcionario", "Periodo", "Base"])
    aba.append(["1001", "Maria", "01/2020", 0.24])
    aba.append(["1001", "Maria", "02/2020", 0.5])
    aba.append(["1001", "Maria", "03/2020", 2.2])
    aba.append(["1001", "Maria", "04/2020", 0.18])
    aba["D3"].number_format = "0.00"
    aba["D4"].number_format = "0.00"

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    valores = {item.competencia: str(item.valor) for serie in preview.valid_records[0].historicos for item in serie.valores}
    assert valores["02/2020"] == "0.50"
    assert valores["03/2020"] == "2.20"


def test_build_preview_keeps_dense_single_decimal_history_values() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["Matricula", "Nome", "CPF", "Admissao", "Demissao"])
    controle.append(["1001", "Maria", "12345678901", "01/01/2020", "02/01/2020"])

    historico = Workbook()
    aba = historico.active
    aba.title = "Maria"
    aba.append(["Matricula", "Funcionário", "Dt.Admissão", "Dt.Demissão", "Período", "Base"])
    aba.append(["1001", "Maria", "01/01/2020", "02/01/2020", "01/2020", 1.2])
    aba.append(["1001", "Maria", "01/01/2020", "02/01/2020", "02/2020", 2.3])
    aba.append(["1001", "Maria", "01/01/2020", "02/01/2020", "03/2020", 3.4])

    preview = build_preview(cadastro, history_workbook=historico, limit=None)

    valores = {item.competencia: str(item.valor) for serie in preview.valid_records[0].historicos for item in serie.valores}
    assert valores["01/2020"] == "1.20"
    assert valores["02/2020"] == "2.30"
    assert valores["03/2020"] == "3.40"


def test_build_preview_allows_existing_history_correction_without_cpf() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["1001", "MARIA SOUZA", None, "15/04/2026", "ok"])

    historico = Workbook()
    aba = historico.active
    aba.title = "MARIA SOUZA"
    aba.append(["Funcionario", "Periodo", "Base"])
    aba.append(["MARIA SOUZA", "01/2020", "10,00"])

    preview = build_preview(
        cadastro,
        history_workbook=historico,
        limit=None,
        execution_mode=ExecutionMode.CORRIGIR_HISTORICO,
    )

    assert len(preview.valid_records) == 1
    assert preview.valid_records[0].cpf == ""
    assert preview.valid_records[0].processo is None
    assert preview.valid_records[0].data_demissao == "15/04/2026"
    assert preview.valid_records[0].historicos[0].nome == "Base"


def test_build_preview_accepts_explicit_calculation_date_for_existing_date_fix_mode() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["1001", "MARIA SOUZA", None, "14/12/2019", "15/04/2026", "ok"])

    historico = Workbook()
    aba = historico.active
    aba.title = "MARIA SOUZA"
    aba.append(["Funcionario", "Periodo", "Base"])
    aba.append(["MARIA SOUZA", "01/2020", "10,00"])

    preview = build_preview(
        cadastro,
        history_workbook=historico,
        limit=None,
        execution_mode=ExecutionMode.CORRIGIR_DATAS_E_HISTORICO,
    )

    assert len(preview.valid_records) == 1
    assert preview.valid_records[0].data_demissao == "14/12/2019"
    assert preview.valid_records[0].data_calculo == "15/04/2026"


def test_build_preview_applies_fixed_process_when_row_does_not_have_one() -> None:
    cadastro = Workbook()
    controle = cadastro.active
    controle.title = "Controle"
    controle.append(["1001", "MARIA SOUZA", None, "15/04/2026", "ok"])

    historico = Workbook()
    aba = historico.active
    aba.title = "MARIA SOUZA"
    aba.append(["Funcionario", "Periodo", "Base"])
    aba.append(["MARIA SOUZA", "01/2020", "10,00"])

    preview = build_preview(
        cadastro,
        history_workbook=historico,
        limit=None,
        execution_mode=ExecutionMode.CORRIGIR_HISTORICO,
        fixed_process="0010953-19.2017.5.03.0034",
    )

    assert len(preview.valid_records) == 1
    assert preview.valid_records[0].processo == "0010953-19.2017.5.03.0034"
