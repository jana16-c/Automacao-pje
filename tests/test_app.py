from datetime import datetime
from threading import Event

from openpyxl import load_workbook
from selenium.common.exceptions import TimeoutException

from pje_automation.app import Application
from pje_automation.domain.execution import ExecutionMode
from pje_automation.domain.errors import AutomationCancelledError, WorkflowExecutionError
from pje_automation.domain.models import HistoricalSeries, HistoricalValue, JobRecord, Record, RecordSource, WorkbookPreview
from pje_automation.domain.states import JobState


def build_application() -> Application:
    app = Application.__new__(Application)
    app.browser_manager = type(
        "BrowserManagerStub",
        (),
        {
            "config": {
                "execution": {
                    "retry_backoff_seconds": 4,
                    "test_mode_first_record_only": False,
                    "max_retries_per_step": 1,
                    "max_retries_pje_server_error": 3,
                    "resume_recent_calculation": True,
                    "resume_previous_run": True,
                    "overwrite_valid_outputs": False,
                    "cooldown_every_records": 5,
                    "cooldown_seconds": 120,
                    "failed_record_revisit_rounds": 1,
                    "record_watchdog_timeout_seconds": 240,
                    "watchdog_poll_seconds": 5,
                }
            }
        },
    )()
    app._cancel_event = Event()
    app._active_driver = None
    return app


def build_record(record_id: str, has_history: bool) -> Record:
    historicos = []
    if has_history:
        historicos = [
            HistoricalSeries(
                nome="BASE",
                valores=[HistoricalValue(competencia="01/2020", valor=1)],
            )
        ]
    return Record(
        record_id=record_id,
        nome=f"Nome {record_id}",
        cpf=record_id.zfill(11),
        data_admissao=None,
        data_demissao="01/01/2020",
        data_calculo=None,
        processo=None,
        historicos=historicos,
        source=RecordSource(sheet="Controle", row=2),
    )


def test_should_retry_workflow_accepts_field_persistence_errors() -> None:
    app = build_application()

    should_retry = app._should_retry_workflow(
        WorkflowExecutionError("FIELD_NOT_PERSISTED", "falhou"),
        attempt=1,
        max_attempts=2,
    )

    assert should_retry is True


def test_pje_server_error_uses_extra_retry_attempts() -> None:
    app = build_application()

    assert app._max_attempts_for_error(WorkflowExecutionError("PJE_SERVER_ERROR", "falhou")) == 4
    assert app._should_retry_workflow(
        WorkflowExecutionError("PJE_SERVER_ERROR", "falhou"),
        attempt=2,
        max_attempts=app._max_attempts_for_error(WorkflowExecutionError("PJE_SERVER_ERROR", "falhou")),
    )
    assert not app._should_retry_workflow(
        WorkflowExecutionError("PJE_SERVER_ERROR", "falhou"),
        attempt=4,
        max_attempts=app._max_attempts_for_error(WorkflowExecutionError("PJE_SERVER_ERROR", "falhou")),
    )


def test_regular_retry_errors_keep_default_attempts() -> None:
    app = build_application()

    assert app._max_attempts_for_error(WorkflowExecutionError("FIELD_NOT_PERSISTED", "falhou")) == 2


def test_resume_recent_calculation_enabled_reads_execution_config() -> None:
    app = build_application()

    assert app._resume_recent_calculation_enabled() is True


def test_write_failure_report_creates_excel_with_error_rows(tmp_path) -> None:
    app = build_application()
    record = build_record("32806201551", True)
    job = JobRecord(
        record_id=record.record_id,
        nome=record.nome,
        cpf_masked="***.***.***-**",
        state=JobState.ERRO,
        updated_at=datetime.now(),
        resume_state=JobState.REGERANDO_FGTS,
        attempt_count=3,
        error_code="PJE_SERVER_ERROR",
        error_message="O PJe retornou erro interno.",
        error_details="url=http://localhost/pjecalc",
    )
    output = app._write_failure_report(
        tmp_path / "falhas.xlsx",
        [(record, WorkflowExecutionError("PJE_SERVER_ERROR", "O PJe retornou erro interno."), job, None)],
    )

    workbook = load_workbook(output)
    sheet = workbook["Falhas"]

    assert sheet.cell(row=1, column=1).value == "Registro"
    assert sheet.cell(row=2, column=1).value == "32806201551"
    assert sheet.cell(row=2, column=6).value == "PJE_SERVER_ERROR"
    assert sheet.cell(row=2, column=8).value == "ERRO"
    assert sheet.cell(row=2, column=9).value == "REGERANDO_FGTS"
    assert sheet.cell(row=2, column=10).value == 3
    assert sheet.cell(row=2, column=11).value == "url=http://localhost/pjecalc"


def test_write_failure_report_uses_timestamp_when_file_is_locked(tmp_path, monkeypatch) -> None:
    import pje_automation.app as app_module

    app = build_application()
    original_save = app_module.Workbook.save
    calls = []

    def fake_save(self, filename):
        calls.append(filename)
        if len(calls) == 1:
            raise PermissionError("locked")
        return original_save(self, filename)

    monkeypatch.setattr(app_module.Workbook, "save", fake_save)

    output = app._write_failure_report(
        tmp_path / "falhas.xlsx",
        [(build_record("32806201551", True), WorkflowExecutionError("PJE_SERVER_ERROR", "falhou"), None, None)],
    )

    assert output.name.startswith("falhas_")
    assert output.exists()


def test_retry_backoff_seconds_reads_execution_config() -> None:
    app = build_application()

    assert app._retry_backoff_seconds() == 4


def test_record_watchdog_timeout_seconds_reads_execution_config() -> None:
    app = build_application()

    assert app._record_watchdog_timeout_seconds() == 240


def test_failed_record_revisit_rounds_reads_execution_config() -> None:
    app = build_application()

    assert app._failed_record_revisit_rounds() == 1


def test_cooldown_profile_keeps_default_pause_for_new_calculation() -> None:
    app = build_application()

    assert app._cooldown_profile(ExecutionMode.NOVO_CALCULO) == (5, 120)


def test_cooldown_profile_disables_default_pause_for_correction_modes() -> None:
    app = build_application()

    assert app._cooldown_profile(ExecutionMode.CORRIGIR_HISTORICO) == (0, 0)
    assert app._cooldown_profile(ExecutionMode.CORRIGIR_DATAS_E_HISTORICO) == (0, 0)


def test_estimate_remaining_seconds_uses_average_of_completed_records() -> None:
    app = build_application()

    remaining = app._estimate_remaining_seconds(handled_records=5, total_records=30, elapsed_seconds=100)

    assert remaining == 500


def test_estimate_remaining_seconds_returns_none_without_completed_records() -> None:
    app = build_application()

    assert app._estimate_remaining_seconds(handled_records=0, total_records=30, elapsed_seconds=100) is None
    assert app._estimate_remaining_seconds(handled_records=30, total_records=30, elapsed_seconds=100) is None


def test_should_retry_workflow_accepts_timeout_exception() -> None:
    app = build_application()

    assert app._should_retry_workflow(TimeoutException("travou"), attempt=1, max_attempts=2) is True


def test_select_records_for_execution_prefers_history_matches() -> None:
    app = build_application()
    preview = WorkbookPreview(
        valid_records=[build_record("1", False), build_record("2", True), build_record("3", False)],
        invalid_rows=[],
        ambiguous_rows=[],
        sheet_names=["Controle"],
    )

    selected = app._select_records_for_execution(preview, history_required=True, apply_test_mode_limit=False)

    assert [record.record_id for record in selected] == ["2"]


def test_select_records_for_execution_raises_when_history_file_has_no_match() -> None:
    app = build_application()
    preview = WorkbookPreview(
        valid_records=[build_record("1", False), build_record("2", False)],
        invalid_rows=[],
        ambiguous_rows=[],
        sheet_names=["Controle"],
    )

    try:
        app._select_records_for_execution(preview, history_required=True, apply_test_mode_limit=False)
    except ValueError as exc:
        assert "historico salarial" in str(exc)
    else:
        raise AssertionError("Era esperado erro de historico sem correspondencia.")


def test_select_records_for_execution_honors_test_mode_limit() -> None:
    app = build_application()
    app.browser_manager.config["execution"]["test_mode_first_record_only"] = True
    preview = WorkbookPreview(
        valid_records=[build_record("1", True), build_record("2", True)],
        invalid_rows=[],
        ambiguous_rows=[],
        sheet_names=["Controle"],
    )

    selected = app._select_records_for_execution(preview, history_required=True, apply_test_mode_limit=True)

    assert [record.record_id for record in selected] == ["1"]


def test_select_records_for_execution_keeps_records_without_history_in_import_mode() -> None:
    app = build_application()
    preview = WorkbookPreview(
        valid_records=[build_record("1", False), build_record("2", False)],
        invalid_rows=[],
        ambiguous_rows=[],
        sheet_names=["Controle"],
    )

    selected = app._select_records_for_execution(preview, history_required=False, apply_test_mode_limit=False)

    assert [record.record_id for record in selected] == ["1", "2"]


def test_build_execution_plan_skips_completed_record_with_outputs(tmp_path) -> None:
    app = build_application()
    record = build_record("32806201551", True)
    pdf_dir = tmp_path / "PDF"
    pjc_dir = tmp_path / "PJC"
    pdf_dir.mkdir()
    pjc_dir.mkdir()
    pdf_file, pjc_file = app._record_output_paths(record, pdf_dir, pjc_dir)
    pdf_file.write_bytes(b"%PDF")
    pjc_file.write_text("PJC")
    repository = type(
        "RepositoryStub",
        (),
        {
            "list_jobs": lambda self: [
                JobRecord(
                    record_id=record.record_id,
                    nome=record.nome,
                    cpf_masked="***",
                    state=JobState.CONCLUIDO,
                    updated_at=datetime.now(),
                )
            ]
        },
    )()
    logger = type("LoggerStub", (), {"info": lambda *args, **kwargs: None})()

    plan, skipped = app._build_execution_plan([record], repository, pdf_dir, pjc_dir, logger)

    assert plan == []
    assert skipped == [record]


def test_build_execution_plan_marks_incomplete_record_for_resume(tmp_path) -> None:
    app = build_application()
    record = build_record("32806201551", True)
    pdf_dir = tmp_path / "PDF"
    pjc_dir = tmp_path / "PJC"
    pdf_dir.mkdir()
    pjc_dir.mkdir()
    repository = type(
        "RepositoryStub",
        (),
        {
            "list_jobs": lambda self: [
                JobRecord(
                    record_id=record.record_id,
                    nome=record.nome,
                    cpf_masked="***",
                    state=JobState.REGERANDO_FGTS,
                    updated_at=datetime.now(),
                )
            ]
        },
    )()
    logger = type("LoggerStub", (), {"info": lambda *args, **kwargs: None})()

    plan, skipped = app._build_execution_plan([record], repository, pdf_dir, pjc_dir, logger)

    assert plan == [(record, JobState.REGERANDO_FGTS)]
    assert skipped == []


def test_resume_state_from_job_prefers_checkpoint_state() -> None:
    app = build_application()
    job = JobRecord(
        record_id="1",
        nome="Teste",
        cpf_masked="***",
        state=JobState.ERRO,
        updated_at=datetime.now(),
        resume_state=JobState.PREENCHENDO_HISTORICO,
    )

    assert app._resume_state_from_job(job) == JobState.PREENCHENDO_HISTORICO


def test_build_revisit_plan_uses_latest_resume_state_from_failed_jobs() -> None:
    app = build_application()
    record = build_record("32806201551", True)
    failed = [
        (
            record,
            WorkflowExecutionError("PJE_SERVER_ERROR", "falhou"),
            JobRecord(
                record_id=record.record_id,
                nome=record.nome,
                cpf_masked="***",
                state=JobState.ERRO,
                updated_at=datetime.now(),
                resume_state=JobState.PREENCHENDO_HISTORICO,
            ),
            None,
        )
    ]

    assert app._build_revisit_plan(failed) == [(record, JobState.PREENCHENDO_HISTORICO)]


def test_request_stop_marks_application_as_cancelled() -> None:
    app = build_application()

    app.request_stop()

    assert app.is_stop_requested() is True


def test_ensure_not_cancelled_raises_after_stop_request() -> None:
    app = build_application()
    app.request_stop()

    try:
        app._ensure_not_cancelled()
    except AutomationCancelledError as exc:
        assert exc.code == "AUTOMATION_CANCELLED"
    else:
        raise AssertionError("Era esperado erro de cancelamento.")
